"""DOL OFLC H-1B/LCA disclosure data source module. Spec: the source-collection validation notes.

The Office of Foreign Labor Certification (OFLC) publishes quarterly Excel workbooks of every
LCA (Labor Condition Application, the filing behind H-1B/H-1B1/E-3 petitions) it disposed of
that quarter — U.S. government public domain, same provenance story as EDGAR (docs/03).

Portal (live-checked 2026-07-03, 200 OK): https://www.dol.gov/agencies/eta/foreign-labor/performance
That HTML page is rendered server-side (no JS needed) and lists a row per program (LCA, PERM,
H-2A, H-2B, CW-1, Prevailing Wage) with a direct .xlsx href for the current fiscal-year quarter,
plus older/legacy files can be reached by pattern once one is known.

Real current-quarter link, confirmed live (curl 200, content-length 137758384 bytes,
last-modified 2026-05-14):
    https://www.dol.gov/media/LCA_Dislclosure_Data_FY2026_Q2.xlsx
Note the DOL's own typo in the path ("Dislclosure_Data", not "Disclosure_Data") — this is not a
guess, it's the literal live href text scraped off the page. Do not "fix" it; DOL owns the file
name and has kept the typo across quarters.

File format (live-verified via raw zip/deflate byte inspection, since openpyxl was not available
in this environment — see the parser below, which is openpyxl-first with a documented fallback):
  - Standard OOXML .xlsx, single sheet ("xl/worksheets/sheet1.xml"), dimension A1:CT1039356 on
    the FY2026 Q2 file i.e. ~1.04M data rows x 98 columns.
  - Row 1 is the header row, cells use shared-string refs (t="s"), consistent with a normal
    Excel export — NOT inline strings, so a shared-strings-aware reader (openpyxl, or any
    correct OOXML reader) is required; a naive per-row XML grep will not recover header text.
  - Column names are DOL's long-stable public LCA disclosure schema (unchanged in shape across
    recent fiscal years per DOL's published record layouts and third-party trackers): among
    others CASE_NUMBER, CASE_STATUS, DECISION_DATE, VISA_CLASS, EMPLOYER_NAME, JOB_TITLE,
    WORKSITE_CITY, WORKSITE_STATE, WAGE_RATE_OF_PAY_FROM, WAGE_RATE_OF_PAY_TO, WAGE_UNIT_OF_PAY.
    Exact per-quarter column names are read from row 1 at parse time rather than hardcoded
    indices, precisely because DOL revises column names/order slightly release to release.

Dependency: requires `openpyxl` (not yet in pyproject.toml — needs to be added there; NOT
installed by this module). If openpyxl is unavailable, fetch_new() logs one clear message and
returns ([], state) unchanged, same "never take down another source" contract as fred.py.

Full-file parsing note: the live FY2026 Q2 file is ~1.04M rows. This module parses the whole
sheet (openpyxl's read_only+iter_rows mode streams it, so a full pass is memory-safe), but a
full run was NOT executed end-to-end in this build pass (no openpyxl available in this sandbox)
— the parser is written and unit-tested against a small hand-built real-schema fixture instead.
Expect a full-quarter parse to take on the order of tens of seconds to a few minutes (~1M rows,
iter_rows streaming) the first time it actually runs live; that cost is paid once per new
quarterly file, not per poll.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime, timezone
from io import BytesIO

import httpx

SOURCE_ID = "dol-h1b-lca"
LABEL = "DOL OFLC H-1B/LCA disclosure data (U.S. government public domain)"

PERFORMANCE_PAGE = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
USER_AGENT = "The Junkyard (botfeeder.junkyard.guru) - contact TBD"

# The live href text has DOL's own typo baked in ("Dislclosure", not "Disclosure") — this
# regex matches the real pattern scraped off dol.gov, not a guess.
_LCA_LINK_RE = re.compile(
    r'href="(https://www\.dol\.gov/media/LCA_Dis[lc]+losure_Data_FY(\d{4})(?:_Q(\d))?\.xlsx)"',
    re.IGNORECASE,
)

# Record fields we normalize out of whatever the real header row names are this quarter.
# Matched case-insensitively / substring-tolerant since DOL revises header text slightly
# release to release (e.g. WAGE_RATE_OF_PAY_FROM vs PW_WAGE_LEVEL naming drift historically).
_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "case_number": ("CASE_NUMBER",),
    "case_status": ("CASE_STATUS",),
    "decision_date": ("DECISION_DATE",),
    "visa_class": ("VISA_CLASS",),
    "employer_name": ("EMPLOYER_NAME",),
    "job_title": ("JOB_TITLE",),
    "worksite_city": ("WORKSITE_CITY",),
    "worksite_state": ("WORKSITE_STATE",),
    "wage_rate_from": ("WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_FROM_1"),
    "wage_rate_to": ("WAGE_RATE_OF_PAY_TO", "WAGE_RATE_OF_PAY_TO_1"),
    "wage_unit": ("WAGE_UNIT_OF_PAY",),
}


def client() -> httpx.Client:
    return httpx.Client(headers={"User-Agent": USER_AGENT}, timeout=60.0)


def find_current_lca_file(html: str) -> dict | None:
    """Parse the performance page HTML for the current-quarter LCA disclosure xlsx link.

    Returns {"url", "fiscal_year", "quarter"} for the first (most prominent) match, which on
    the live page is the current-quarter row; falls back to the newest FY/Q found if several
    are present, since the page also links prior-quarter files further down.
    """
    matches = [
        {"url": m.group(1), "fiscal_year": int(m.group(2)), "quarter": int(m.group(3)) if m.group(3) else 0}
        for m in _LCA_LINK_RE.finditer(html)
    ]
    if not matches:
        return None
    matches.sort(key=lambda m: (m["fiscal_year"], m["quarter"]), reverse=True)
    return matches[0]


def _header_index(headers: list[str]) -> dict[str, int]:
    upper = [(h or "").strip().upper() for h in headers]
    out: dict[str, int] = {}
    for field, aliases in _FIELD_ALIASES.items():
        for alias in aliases:
            if alias in upper:
                out[field] = upper.index(alias)
                break
    return out


def parse_lca_workbook(content: bytes, *, source_url: str, fetched_at: str) -> list[dict]:
    """Parse one LCA disclosure .xlsx into normalized LCA records.

    Streams via openpyxl's read_only + iter_rows so a ~1M-row workbook doesn't need to be
    materialized in memory. Raises ImportError if openpyxl isn't installed — caller decides
    whether that's fatal (fetch_new() treats it as "skip this cycle", per fred.py's pattern).
    """
    import openpyxl  # deferred import: optional dependency, see module docstring

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return []
    idx = _header_index(headers)

    records: list[dict] = []
    for row in rows:
        def get(field: str):
            i = idx.get(field)
            return row[i] if i is not None and i < len(row) else None

        case_number = get("case_number")
        if not case_number:
            continue
        decision_date = get("decision_date")
        if hasattr(decision_date, "isoformat"):
            decision_date = decision_date.isoformat()
        records.append({
            "case_number": case_number,
            "case_status": get("case_status"),
            "decision_date": decision_date,
            "visa_class": get("visa_class"),
            "employer_name": get("employer_name"),
            "job_title": get("job_title"),
            "worksite_city": get("worksite_city"),
            "worksite_state": get("worksite_state"),
            "wage_rate_from": get("wage_rate_from"),
            "wage_rate_to": get("wage_rate_to"),
            "wage_unit": get("wage_unit"),
            "source_url": source_url,
            "fetched_at": fetched_at,
        })
    return records


_warned = False


def _warn_once(msg: str) -> None:
    global _warned
    if not _warned:
        print(f"[producer:{SOURCE_ID}] {msg}", file=sys.stderr)
        _warned = True


def fetch_new(state: dict, c: httpx.Client) -> tuple[list[dict], dict]:
    """One poll cycle: check the performance page for a newer quarterly file than the one
    recorded in state; if found, download + parse it in full. Cadence: this is a quarterly
    source — poll daily is fine (the HEAD-equivalent page fetch is cheap; the multi-hundred-MB
    download only happens when a new quarter is actually detected)."""
    now = datetime.now(timezone.utc).isoformat()
    last_url = state.get("last_file_url")

    try:
        r = c.get(PERFORMANCE_PAGE)
        r.raise_for_status()
    except Exception as e:  # noqa: BLE001 — one bad cycle must not stop other sources
        print(f"[producer:{SOURCE_ID}] fetch failed: {e}", file=sys.stderr)
        return [], state

    current = find_current_lca_file(r.text)
    if not current:
        _warn_once("no LCA disclosure link found on performance page — page layout may have changed")
        return [], state

    if current["url"] == last_url:
        return [], state  # already processed this quarter's file

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        _warn_once(
            "openpyxl not installed — skipping DOL H-1B/LCA parse. "
            "Add 'openpyxl' to pyproject.toml dependencies to enable this source."
        )
        return [], state

    try:
        dl = c.get(current["url"])
        dl.raise_for_status()
        records = parse_lca_workbook(dl.content, source_url=current["url"], fetched_at=now)
    except Exception as e:  # noqa: BLE001
        print(f"[producer:{SOURCE_ID}] download/parse failed for {current['url']}: {e}", file=sys.stderr)
        return [], state

    state["last_file_url"] = current["url"]
    state["last_fiscal_year"] = current["fiscal_year"]
    state["last_quarter"] = current["quarter"]
    return records, state
